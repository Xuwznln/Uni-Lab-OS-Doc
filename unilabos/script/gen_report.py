#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
场景二实验报告一键生成脚本(三合一).

一个脚本, 一个输出文件(默认 data/实验报告.xlsx), 含以下 sheet:
  1) RNA浓度检测      —— 来源: BioTek Synergy H1 荧光 CSV + 96 孔板布局
  2) 板排布           —— 来源: LightCycler qPCR XML + 96 孔板布局(96->384 映射)
  3) 扩增曲线图        —— 同上, 全部 384 孔全程荧光曲线(嵌图)
  4) 曲线数据          —— 同上, 每孔逐采集点原始荧光
  5) qPCR统计与分析    —— 阈值法 CT + ΔΔCT(Excel 公式)

数据源(均为真实读取, 仅"荧光->浓度"标曲与 CT 阈值为可调参数):
  --fluor-csv  BioTek 荧光 CSV(RNA 浓度)
  --xml        qPCR 原始数据 XML
  --plate-map  96 孔细胞培养板布局 xlsx(样本身份, 两部分共用)

注意:
  - RNA 浓度换算用占位线性标曲 conc=slope*(RFU-空白)+intercept(默认 slope=1/intercept=0),
    拿到真实标曲后用 --slope/--intercept 替换。荧光法不产生 A260/A280、A260/A230。
  - 96->384 映射(用户确认): 每个 96 孔 -> 384 的 2x2 区块, 左列(奇)=target/目的基因,
    右列(偶)=reference/内参基因, 上下两行=复孔 1/2。
  - CT 用阈值法(默认阈值 3): 扩增段曲线从下往上首次穿过阈值线的循环号(线性插值)。

用法:
  python3 gen_report.py
  python3 gen_report.py --slope 0.01 --intercept 0 --threshold 3 --out data/实验报告.xlsx
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import statistics
import xml.etree.ElementTree as ET

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
from matplotlib.patches import Rectangle  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.drawing.image import Image as XLImage  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# 中文字体：跨平台列表，matplotlib 取第一个已安装的。
# 工作站是 Windows(SimHei/微软雅黑自带)，开发机多为 macOS，CI/Linux 兜底 Noto/文泉驿。
_CJK_FONT_CANDIDATES = [
    "Microsoft YaHei", "SimHei", "SimSun", "Microsoft JhengHei",  # Windows
    "Arial Unicode MS", "Hiragino Sans GB", "Heiti TC", "STHeiti", "Songti SC",  # macOS
    "Noto Sans CJK SC", "Source Han Sans SC", "Source Han Sans CN",
    "WenQuanYi Zen Hei", "WenQuanYi Micro Hei", "Noto Sans CJK JP",  # Linux
]


def _setup_cjk_font() -> None:
    """把首个已安装的中文字体设为 matplotlib 默认 sans-serif，避免中文显示为空心方块(□)。

    matplotlib 仅在字体已注册时才生效；这里显式查 ``fontManager`` 命中后置顶，
    并保留整列表作为回退。命中不到时记一次告警(图中文会变 □，但不阻断出图)。
    """
    import matplotlib.font_manager as fm

    installed = {f.name for f in fm.fontManager.ttflist}
    chosen = next((name for name in _CJK_FONT_CANDIDATES if name in installed), None)
    ordered = ([chosen] if chosen else []) + _CJK_FONT_CANDIDATES + ["DejaVu Sans"]
    matplotlib.rcParams["font.sans-serif"] = ordered
    matplotlib.rcParams["font.family"] = "sans-serif"
    matplotlib.rcParams["axes.unicode_minus"] = False
    if chosen is None:
        import logging

        logging.getLogger(__name__).warning(
            "[gen_report] 未找到中文字体(候选: %s)，图中中文可能显示为方块；"
            "请在运行机(工作站)安装 SimHei/微软雅黑 等中文字体。",
            ", ".join(_CJK_FONT_CANDIDATES),
        )


_setup_cjk_font()

ROWS8 = "ABCDEFGH"             # 96 孔 8 行
ROWS384 = "ABCDEFGHIJKLMNOP"   # 384 孔 16 行
COLS384 = list(range(1, 25))   # 384 孔 24 列
TARGET_COLOR = "#d62728"       # 目的基因
REF_COLOR = "#1f77b4"          # 内参基因

_HERE = os.path.dirname(os.path.abspath(__file__))
_REPO = os.path.abspath(os.path.join(_HERE, "..", ".."))
DEFAULT_FLUOR = os.path.join(_REPO, "data", "test0420_260604_124254.csv")
DEFAULT_XML = os.path.join(_REPO, "data", "场景二实验数据-qpcr20260520181417.xml")
DEFAULT_MAP = os.path.join(_REPO, "data", "实验二细胞培养板.xlsx")
DEFAULT_OUT = os.path.join(_REPO, "data", "实验报告.xlsx")


# ===========================================================================
# 公共: 读 96 孔板布局
# ===========================================================================

def read_plate_map(xlsx_path: str):
    """读 96 孔板布局 xlsx, 返回 {(R,C): label}, R 行 1..8, C 列 1..12."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    map96: dict[tuple[int, int], str] = {}
    for r in range(1, 9):
        for c in range(1, 13):
            cell = ws.cell(row=r + 1, column=c + 1).value
            if cell is not None and str(cell).strip():
                map96[(r, c)] = str(cell).strip()
    return map96


def is_blank(label: str) -> bool:
    return "空白" in label


def sample_sort_key(label: str):
    if is_blank(label):
        return (-1, 0)
    m = re.search(r"(\d+)", label)
    return (0, int(m.group(1)) if m else 999)


# ===========================================================================
# 第 1 部分: RNA 浓度检测(BioTek 荧光)
# ===========================================================================

def parse_biotek_csv(csv_path: str):
    """解析 BioTek 荧光 CSV, 返回 (rfu_grid, meta).
    rfu_grid: {(R,C): float}; meta: dict(摘要, 用于注释).
    """
    rows = []
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        for line in csv.reader(f, delimiter="\t"):
            rows.append(line)

    meta = {}
    for r in rows:
        if not r:
            continue
        key = r[0].strip()
        if key in ("Date", "Time", "Reader Type:", "Reader Serial Number:") and len(r) > 1:
            meta[key.rstrip(":")] = r[1].strip()

    grid: dict[tuple[int, int], float] = {}
    in_results = False
    for r in rows:
        if r and r[0].strip() == "Results":
            in_results = True
            continue
        if in_results and r and r[0].strip() in ROWS8:
            R = ROWS8.index(r[0].strip()) + 1
            for C in range(1, 13):
                if C < len(r):
                    cell = r[C].strip()
                    if cell and re.fullmatch(r"-?\d+(\.\d+)?", cell):
                        grid[(R, C)] = float(cell)
    return grid, meta


def fluor_to_concentration(x: float, slope: float, intercept: float) -> float:
    """占位标准曲线: 由(空白扣除后的)荧光换算 RNA 浓度 ng/uL."""
    return slope * x + intercept


def compute_rna(fluor_csv, plate_map_path, slope, intercept, blank_mode):
    """返回 (conc_by_sample, meta, blank_mean)."""
    grid, meta = parse_biotek_csv(fluor_csv)
    map96 = read_plate_map(plate_map_path)

    blank_rfus = [grid[k] for k, lab in map96.items() if is_blank(lab) and k in grid]
    blank_mean = statistics.mean(blank_rfus) if blank_rfus else 0.0

    conc_by_sample: dict[str, list[float]] = {}
    for (R, C), label in map96.items():
        if (R, C) not in grid:
            continue
        x = grid[(R, C)] - blank_mean if blank_mode == "subtract" else grid[(R, C)]
        conc_by_sample.setdefault(label, []).append(fluor_to_concentration(x, slope, intercept))
    return conc_by_sample, meta, blank_mean


def add_rna_sheet(wb, conc_by_sample, fluor_csv, meta, blank_mean, slope, intercept):
    ws = wb.create_sheet("RNA浓度检测")
    bold = Font(bold=True)
    ws["A1"] = "录入者签名及日期"
    ws["A2"] = "复核人签名及日期"
    ws["A3"] = (f"数据来源(荧光): {os.path.basename(fluor_csv)}; "
                f"仪器: {meta.get('Reader Type','')} {meta.get('Date','')}")
    ws["A4"] = (f"换算: 占位线性标曲 conc={slope}*(RFU-空白)+{intercept}; "
                f"空白RFU均值={blank_mean:.2f}; 荧光法不含 A260/A280、A260/A230")

    header = ["Sample Name", "Nucleic Acid(ng/uL)", "n复孔", "SD"]
    hrow = 5
    for j, h in enumerate(header, start=1):
        ws.cell(row=hrow, column=j, value=h).font = bold

    r = hrow + 1
    for lab in sorted(conc_by_sample.keys(), key=sample_sort_key):
        vals = conc_by_sample[lab]
        ws.cell(row=r, column=1, value=lab)
        ws.cell(row=r, column=2, value=round(statistics.mean(vals), 4))
        ws.cell(row=r, column=3, value=len(vals))
        ws.cell(row=r, column=4, value=round(statistics.stdev(vals), 4) if len(vals) > 1 else 0.0)
        r += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    return ws


# ===========================================================================
# 第 2 部分: qPCR 解析 + 96->384 映射 + 绘图
# ===========================================================================

def parse_qpcr_xml(xml_path: str, max_points: int | None = None):
    """单遍 iterparse 解析, 返回 (pos_by_sample, fluor_by_sample).
    max_points 为 None 取全部采集点(扩增+熔解); 否则只取前 max_points 个.
    """
    pos_by_sample: dict[str, str] = {}
    fluor_by_sample: dict[str, list[float]] = {}

    for _ev, el in ET.iterparse(xml_path, events=("end",)):
        if el.tag == "Sample":
            num = el.get("Number")
            if num is not None:
                vals: list[float] = []
                for acq in el.findall("Acq"):
                    if max_points is not None:
                        k = acq.get("Number")
                        if k is None or int(k) > max_points:
                            continue
                    chan = acq.find("Chan")
                    if chan is None:
                        continue
                    fl = chan.find("prop[@name='Fluor']")
                    if fl is not None and fl.text:
                        vals.append(float(fl.text))
                fluor_by_sample[num] = vals
            el.clear()
        elif el.tag == "AnalysisSample":
            name_el = el.find("prop[@name='name']")
            pos_el = el.find("prop[@name='Position']")
            if name_el is not None and name_el.text and pos_el is not None and pos_el.text:
                sample_no = name_el.text.strip().split()[-1]
                pos_by_sample[sample_no] = pos_el.text.strip()
            el.clear()

    return pos_by_sample, fluor_by_sample


def parse_qpcr_melt(xml_path: str, amp_cycles: int = 45):
    """解析熔解段, 返回 (pos_by_sample, melt_by_sample)。

    单遍 iterparse: 取每孔扩增循环之后(Acq 序号 > amp_cycles)的采集点作为熔解段,
    每点保留 (Temp, Fluor); 同时收集 AnalysisSample 的 sample_no->Position 映射。
    melt_by_sample[num] = (temps[], fluors[]).
    """
    pos_by_sample: dict[str, str] = {}
    melt_by_sample: dict[str, tuple[list[float], list[float]]] = {}

    for _ev, el in ET.iterparse(xml_path, events=("end",)):
        if el.tag == "Sample":
            num = el.get("Number")
            if num is not None:
                temps: list[float] = []
                fluors: list[float] = []
                for acq in el.findall("Acq"):
                    k = acq.get("Number")
                    if k is None or int(k) <= amp_cycles:
                        continue
                    chan = acq.find("Chan")
                    if chan is None:
                        continue
                    fl = chan.find("prop[@name='Fluor']")
                    tp = chan.find("prop[@name='Temp']")
                    if fl is not None and fl.text and tp is not None and tp.text:
                        temps.append(float(tp.text))
                        fluors.append(float(fl.text))
                if temps:
                    melt_by_sample[num] = (temps, fluors)
            el.clear()
        elif el.tag == "AnalysisSample":
            name_el = el.find("prop[@name='name']")
            pos_el = el.find("prop[@name='Position']")
            if name_el is not None and name_el.text and pos_el is not None and pos_el.text:
                sample_no = name_el.text.strip().split()[-1]
                pos_by_sample[sample_no] = pos_el.text.strip()
            el.clear()

    return pos_by_sample, melt_by_sample


def pos_str(row384: int, col384: int) -> str:
    return f"{ROWS384[row384 - 1]}{col384}"


def build_well_meta(map96):
    """96 孔展开成 384 孔, 返回 {position: meta(sample, gene, rep, src96)}."""
    meta: dict[str, dict] = {}
    for (r, c), label in map96.items():
        tcol, rcol = 2 * c - 1, 2 * c
        top, bot = 2 * r - 1, 2 * r
        meta[pos_str(top, tcol)] = dict(sample=label, gene="target", rep=1, src96=(r, c))
        meta[pos_str(bot, tcol)] = dict(sample=label, gene="target", rep=2, src96=(r, c))
        meta[pos_str(top, rcol)] = dict(sample=label, gene="reference", rep=1, src96=(r, c))
        meta[pos_str(bot, rcol)] = dict(sample=label, gene="reference", rep=2, src96=(r, c))
    return meta


def short_label(label: str) -> str:
    if not label:
        return ""
    if is_blank(label):
        return "空白"
    m = re.search(r"(\d+)", label)
    return m.group(1) if m else label


def sample_color_map(map96):
    labels, seen = [], set()
    for lab in map96.values():
        if lab not in seen:
            seen.add(lab)
            labels.append(lab)
    labels = sorted(labels, key=sample_sort_key)
    cmap = plt.get_cmap("tab20")
    colors = {}
    for i, lab in enumerate(labels):
        colors[lab] = "#d9d9d9" if is_blank(lab) else cmap(i % 20)
    return colors, labels


def plot_plate_layout(meta, map96, out_png, dpi):
    colors, labels = sample_color_map(map96)
    fig, ax = plt.subplots(figsize=(16, 11))
    for pos, m in meta.items():
        row = ROWS384.index(pos[0]) + 1
        col = int(pos[1:])
        x, y = col - 1, 16 - row
        ax.add_patch(Rectangle((x, y), 1, 1, facecolor=colors.get(m["sample"], "#fff"),
                               edgecolor="white", linewidth=0.5))
        gene_c = TARGET_COLOR if m["gene"] == "target" else REF_COLOR
        gtag = "T" if m["gene"] == "target" else "R"
        ax.text(x + 0.5, y + 0.62, short_label(m["sample"]),
                ha="center", va="center", fontsize=5.5, color="black")
        ax.text(x + 0.5, y + 0.28, f"{gtag}{m['rep']}",
                ha="center", va="center", fontsize=5, color=gene_c, fontweight="bold")
    for c in COLS384:
        ax.text(c - 0.5, 16.25, str(c), ha="center", va="bottom", fontsize=7)
    for i, rl in enumerate(ROWS384):
        ax.text(-0.25, 16 - (i + 1) + 0.5, rl, ha="right", va="center", fontsize=7)
    ax.set_xlim(-0.6, 24.2)
    ax.set_ylim(-0.4, 16.7)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("qPCR 板排布 (384 孔)  T=目的基因 target  R=内参基因 reference", fontsize=12)
    from matplotlib.patches import Patch
    handles = [Patch(facecolor=colors[l], edgecolor="white", label=l) for l in labels]
    ax.legend(handles=handles, loc="upper left", bbox_to_anchor=(1.0, 1.0),
              fontsize=7, title="样本", title_fontsize=8)
    fig.tight_layout()
    fig.savefig(out_png, dpi=dpi, bbox_inches="tight")
    plt.close(fig)


def plot_amp_curves(fluor_by_pos, meta, out_png, dpi, amp_cycles=45, logy=False):
    fig, ax = plt.subplots(figsize=(14, 7))
    n_t = n_r = 0
    max_len = 0
    for pos, vals in fluor_by_pos.items():
        if not vals:
            continue
        max_len = max(max_len, len(vals))
        gene = (meta.get(pos) or {}).get("gene")
        if gene == "target":
            color, n_t = TARGET_COLOR, n_t + 1
        elif gene == "reference":
            color, n_r = REF_COLOR, n_r + 1
        else:
            color = "#999999"
        ax.plot(range(1, len(vals) + 1), vals, color=color, lw=0.5, alpha=0.35)
    if 0 < amp_cycles < max_len:
        ax.axvline(amp_cycles + 0.5, color="#444444", ls="--", lw=1)
        ymax = ax.get_ylim()[1]
        ax.text(amp_cycles / 2, ymax, "扩增段 (1-%d)" % amp_cycles,
                ha="center", va="top", fontsize=9, color="#444444")
        ax.text((amp_cycles + max_len) / 2, ymax, "熔解段 (%d-%d)" % (amp_cycles + 1, max_len),
                ha="center", va="top", fontsize=9, color="#444444")
    if logy:
        ax.set_yscale("log")
    from matplotlib.lines import Line2D
    handles = [
        Line2D([0], [0], color=TARGET_COLOR, lw=2, label=f"目的基因 target (n={n_t})"),
        Line2D([0], [0], color=REF_COLOR, lw=2, label=f"内参基因 reference (n={n_r})"),
    ]
    ax.legend(handles=handles, fontsize=9, loc="upper left")
    ax.set_xlabel("采集点序号 (Acquisition index)")
    ax.set_ylabel("荧光 (Fluorescence)" + ("  [log]" if logy else ""))
    ax.set_title("qPCR 全程原始荧光曲线 (全部 384 孔, 每孔 %d 点)" % max_len)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_png, dpi=dpi, bbox_inches="tight")
    plt.close(fig)


def plot_melt_curves(melt_by_pos, meta, out_png, dpi, smooth=9,
                     tmin_margin=1.5, tmax_margin=1.0):
    """熔解曲线: 逐孔按温度排序后求 -dF/dT, 画 -dF/dT vs 温度(全部 384 孔)。

    smooth: 荧光移动平均窗口(奇数, >1 启用), 采用边缘填充避免序列首尾的卷积伪影。
    tmin_margin/tmax_margin: 裁掉斜升起止的热/光学稳定瞬态(默认起始+1.5°C、末端-1.0°C),
        避免边界尖峰压垮 y 轴; 导数仍在完整序列上计算后再按窗口截取。
    """
    import numpy as np

    fig, ax = plt.subplots(figsize=(14, 7))
    n_t = n_r = 0
    y_pool: list[float] = []
    for pos, (temps, fluors) in melt_by_pos.items():
        if not temps or len(temps) < 5:
            continue
        t = np.asarray(temps, dtype=float)
        f = np.asarray(fluors, dtype=float)
        order = np.argsort(t)
        t, f = t[order], f[order]
        if smooth and smooth > 1 and len(f) >= smooth:
            pad = smooth // 2
            fp = np.pad(f, pad, mode="edge")
            f = np.convolve(fp, np.ones(smooth) / smooth, mode="valid")[:len(t)]
        d = -np.gradient(f, t)
        # 截取稳定分析窗口, 去掉斜升起止瞬态
        win = (t >= t.min() + tmin_margin) & (t <= t.max() - tmax_margin)
        if win.sum() < 3:
            win = np.ones_like(t, dtype=bool)
        tw, dw = t[win], d[win]
        gene = (meta.get(pos) or {}).get("gene")
        if gene == "target":
            color, n_t = TARGET_COLOR, n_t + 1
        elif gene == "reference":
            color, n_r = REF_COLOR, n_r + 1
        else:
            color = "#999999"
        ax.plot(tw, dw, color=color, lw=0.5, alpha=0.35)
        y_pool.extend(dw.tolist())

    # 稳健 y 轴范围(1%-99% 分位 + 留白), 防个别孔异常值压垮整图
    if y_pool:
        arr = np.asarray(y_pool, dtype=float)
        lo, hi = np.percentile(arr, 1), np.percentile(arr, 99)
        pad = max((hi - lo) * 0.1, 1e-6)
        ax.set_ylim(lo - pad, hi + pad)

    from matplotlib.lines import Line2D
    handles = [
        Line2D([0], [0], color=TARGET_COLOR, lw=2, label=f"目的基因 target (n={n_t})"),
        Line2D([0], [0], color=REF_COLOR, lw=2, label=f"内参基因 reference (n={n_r})"),
    ]
    ax.legend(handles=handles, fontsize=9, loc="upper left")
    ax.set_xlabel("温度 Temperature (°C)")
    ax.set_ylabel("-dF/dT (Derivative)")
    ax.set_title("qPCR 熔解曲线 Melt Curve (-dF/dT, 全部 384 孔)")
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_png, dpi=dpi, bbox_inches="tight")
    plt.close(fig)


def add_plate_sheets(wb, layout_png, curve_png, fluor_by_pos, meta, npoints):
    ws1 = wb.create_sheet("板排布")
    ws1.cell(row=1, column=1, value="板排布 (单元格: 样本 | T目的/R内参 + 复孔号)")
    for c in COLS384:
        ws1.cell(row=2, column=c + 1, value=c)
    for ri, rl in enumerate(ROWS384):
        ws1.cell(row=ri + 3, column=1, value=rl)
        for c in COLS384:
            m = meta.get(f"{rl}{c}")
            if m:
                gtag = "T" if m["gene"] == "target" else "R"
                ws1.cell(row=ri + 3, column=c + 1, value=f"{m['sample']}|{gtag}{m['rep']}")
    ws1.add_image(XLImage(layout_png), f"A{3 + 16 + 2}")

    ws2 = wb.create_sheet("扩增曲线图")
    ws2.add_image(XLImage(curve_png), "A1")

    ws3 = wb.create_sheet("曲线数据")
    header = ["Position", "96孔源", "样本", "基因", "复孔"] + [f"Acq_{i}" for i in range(1, npoints + 1)]
    ws3.append(header)
    for pos in sorted(fluor_by_pos.keys(), key=lambda p: (ROWS384.index(p[0]), int(p[1:]))):
        m = meta.get(pos, {})
        src = m.get("src96")
        src_s = f"{ROWS384[src[0]-1]}{src[1]}" if src else ""
        gene = {"target": "目的基因", "reference": "内参基因"}.get(m.get("gene", ""), "")
        ws3.append([pos, src_s, m.get("sample", ""), gene, m.get("rep", "")] + fluor_by_pos[pos])


# ===========================================================================
# 第 3 部分: qPCR 统计与分析(阈值法 CT + ΔΔCT)
# ===========================================================================

def ct_threshold(vals, thr):
    """阈值法: 曲线从下往上首次穿过 thr 的循环号(线性插值, 1-based). 无交点返回 None."""
    if not vals:
        return None
    if vals[0] >= thr:
        return 1.0
    for i in range(1, len(vals)):
        y0, y1 = vals[i - 1], vals[i]
        if y0 < thr <= y1:
            return i + (thr - y0) / (y1 - y0)
    return None


def build_rows(map96, ct_by_pos):
    """每个 96 孔展开成 2 个复孔行, 返回按 样本->96孔->复孔 排序的列表."""
    rows = []
    for (r, c), label in map96.items():
        tcol, rcol = 2 * c - 1, 2 * c
        top, bot = 2 * r - 1, 2 * r
        for rep, r384 in ((1, top), (2, bot)):
            tpos, rpos = pos_str(r384, tcol), pos_str(r384, rcol)
            rows.append(dict(
                sample=label,
                pos_label=f"{tpos} {rpos}",
                pos_target=tpos,
                pos_ref=rpos,
                ct_target=ct_by_pos.get(tpos),
                ct_ref=ct_by_pos.get(rpos),
                _sortk=(sample_sort_key(label), r, c, rep),
            ))
    rows.sort(key=lambda d: d["_sortk"])
    return rows


def compute_ave_blank(rows):
    blank_dcts = [
        r["ct_target"] - r["ct_ref"]
        for r in rows
        if is_blank(r["sample"]) and r["ct_ref"] is not None and r["ct_target"] is not None
    ]
    return sum(blank_dcts) / len(blank_dcts) if blank_dcts else None


def add_stats_sheet(wb, rows, xml_path, threshold, ave_blank):
    ws = wb.create_sheet("qPCR统计与分析")
    bold = Font(bold=True)
    ws["A1"] = "录入者签名及日期"
    ws["A2"] = "复核人签名及日期"
    ws["A3"] = (f"数据来源: {os.path.basename(xml_path)}; CT=阈值法(阈值={threshold}); "
                f"Ave.Blank(空白ΔCT均值)={ave_blank if ave_blank is not None else 'NA'}")

    header = ["位置(目的孔 内参孔)", "样本", "内参基因", "目的基因",
              "ΔCT", "Ave.Blank", "ΔΔCT", "2^(-ΔΔCT)", "Mean", "SD"]
    hrow = 4
    for j, h in enumerate(header, start=1):
        ws.cell(row=hrow, column=j, value=h).font = bold

    col = {name: get_column_letter(j) for j, name in enumerate(
        ["pos", "sample", "ref", "target", "dct", "aveblank", "ddct", "rq", "mean", "sd"], start=1)}
    first_data = hrow + 1
    ave_cell = f"${col['aveblank']}${first_data}"

    r = first_data
    group_start = r
    prev_sample = None
    group_bounds = []
    for row in rows:
        if prev_sample is not None and row["sample"] != prev_sample:
            group_bounds.append((prev_sample, group_start, r - 1))
            group_start = r
        prev_sample = row["sample"]

        ws.cell(row=r, column=1, value=row["pos_label"])
        ws.cell(row=r, column=2, value=row["sample"])
        ct_ref, ct_tar = row["ct_ref"], row["ct_target"]
        ws.cell(row=r, column=3, value=round(ct_ref, 4) if ct_ref is not None else "Undetermined")
        ws.cell(row=r, column=4, value=round(ct_tar, 4) if ct_tar is not None else "Undetermined")
        if ct_ref is not None and ct_tar is not None:
            ws.cell(row=r, column=5, value=f"={col['target']}{r}-{col['ref']}{r}")
            ws.cell(row=r, column=7, value=f"={col['dct']}{r}-{ave_cell}")
            ws.cell(row=r, column=8, value=f"=2^(-{col['ddct']}{r})")
        r += 1
    group_bounds.append((prev_sample, group_start, r - 1))

    if ave_blank is not None:
        ws.cell(row=first_data, column=6, value=round(ave_blank, 4))

    for _sample, a, b in group_bounds:
        rng = f"{col['rq']}{a}:{col['rq']}{b}"
        ws.cell(row=a, column=9, value=f"=IFERROR(AVERAGE({rng}),\"\")")
        ws.cell(row=a, column=10, value=f"=IFERROR(STDEV({rng}),\"\")")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    return ws


# ===========================================================================
# 可 import 封装：供节点 m(compute_experiment2_result) 直接调用
# ===========================================================================

def build_rna_table(
    fluor_csv: str,
    plate_map_path: str = DEFAULT_MAP,
    slope: float = 1.0,
    intercept: float = 0.0,
    blank_mode: str = "subtract",
) -> dict:
    """从 BioTek 荧光 CSV 生成 RNA 浓度检测结构化表格(供记录本原生 table 节点)。

    Returns:
        {"header": [...列名...], "rows": [[单元格字符串, ...], ...], "meta": {...}}
    """
    conc_by_sample, meta, blank_mean = compute_rna(
        fluor_csv, plate_map_path, slope, intercept, blank_mode
    )
    header = ["Sample Name", "Nucleic Acid(ng/uL)", "n复孔", "SD"]
    rows: list[list[str]] = []
    for lab in sorted(conc_by_sample.keys(), key=sample_sort_key):
        vals = conc_by_sample[lab]
        mean = round(statistics.mean(vals), 4)
        sd = round(statistics.stdev(vals), 4) if len(vals) > 1 else 0.0
        rows.append([str(lab), f"{mean}", str(len(vals)), f"{sd}"])
    return {
        "header": header,
        "rows": rows,
        "meta": {
            "source": os.path.basename(fluor_csv),
            "reader": meta.get("Reader Type", ""),
            "date": meta.get("Date", ""),
            "blank_mean": round(blank_mean, 2),
            "slope": slope,
            "intercept": intercept,
        },
    }


def render_qpcr_curve_image(
    xml_path: str,
    out_png: str,
    plate_map_path: str = DEFAULT_MAP,
    amp_cycles: int = 45,
    max_points: int | None = None,
    dpi: int = 150,
    logy: bool = False,
) -> str:
    """从 qPCR XML 生成扩增曲线 PNG(供记录本 img 节点)，返回 out_png 路径。"""
    pos_by_sample, fluor_by_sample = parse_qpcr_xml(xml_path, max_points)
    fluor_by_pos = {
        pos_by_sample[sn]: vals
        for sn, vals in fluor_by_sample.items()
        if sn in pos_by_sample
    }
    map96 = read_plate_map(plate_map_path)
    meta = build_well_meta(map96)
    out_dir = os.path.dirname(os.path.abspath(out_png))
    os.makedirs(out_dir, exist_ok=True)
    plot_amp_curves(fluor_by_pos, meta, out_png, dpi, amp_cycles=amp_cycles, logy=logy)
    return out_png


def render_qpcr_melt_image(
    xml_path: str,
    out_png: str,
    plate_map_path: str = DEFAULT_MAP,
    amp_cycles: int = 45,
    dpi: int = 150,
) -> str:
    """从 qPCR XML 熔解段生成熔解曲线 PNG(-dF/dT vs 温度, 供记录本 img 节点)，返回 out_png 路径。"""
    pos_by_sample, melt_by_sample = parse_qpcr_melt(xml_path, amp_cycles=amp_cycles)
    melt_by_pos = {
        pos_by_sample[sn]: tf
        for sn, tf in melt_by_sample.items()
        if sn in pos_by_sample
    }
    map96 = read_plate_map(plate_map_path)
    meta = build_well_meta(map96)
    out_dir = os.path.dirname(os.path.abspath(out_png))
    os.makedirs(out_dir, exist_ok=True)
    plot_melt_curves(melt_by_pos, meta, out_png, dpi)
    return out_png


def render_qpcr_plate_layout_image(
    xml_path: str,
    out_png: str,
    plate_map_path: str = DEFAULT_MAP,
    dpi: int = 150,
) -> str:
    """从 96 孔布局生成 384 孔板排布 PNG(可选，供记录本 img 节点)，返回 out_png 路径。"""
    map96 = read_plate_map(plate_map_path)
    meta = build_well_meta(map96)
    out_dir = os.path.dirname(os.path.abspath(out_png))
    os.makedirs(out_dir, exist_ok=True)
    plot_plate_layout(meta, map96, out_png, dpi)
    return out_png


def build_qpcr_stats_table(
    xml_path: str,
    plate_map_path: str = DEFAULT_MAP,
    threshold: float = 3.0,
    amp_cycles: int = 45,
    max_points: int | None = None,
) -> dict:
    """从 qPCR XML 生成 ΔΔCT 统计分析结构化表格(供记录本原生 table 节点)。

    阈值法 CT → ΔCT=CT目的-CT内参 → ΔΔCT=ΔCT-Ave.Blank(空白ΔCT均值)
    → 相对定量 RQ=2^(-ΔΔCT)，按样本聚合 RQ 均值/标准差(同组仅首行显示均值/SD)。

    Returns:
        {"header": [...], "rows": [[单元格字符串...], ...], "meta": {...}}
    """
    pos_by_sample, fluor_by_sample = parse_qpcr_xml(xml_path, max_points)
    fluor_by_pos = {
        pos_by_sample[sn]: vals
        for sn, vals in fluor_by_sample.items()
        if sn in pos_by_sample
    }
    map96 = read_plate_map(plate_map_path)
    ct_by_pos = {
        pos: ct_threshold(vals[:amp_cycles], threshold)
        for pos, vals in fluor_by_pos.items()
    }
    rows_raw = build_rows(map96, ct_by_pos)
    ave_blank = compute_ave_blank(rows_raw)

    def _num(v) -> str:
        return "" if v is None else f"{round(v, 4)}"

    # 逐复孔算 ΔCT/ΔΔCT/RQ，并按样本收集 RQ 以求均值/SD
    enriched = []
    rq_by_sample: dict[str, list[float]] = {}
    for r in rows_raw:
        ct_t, ct_r = r["ct_target"], r["ct_ref"]
        dct = ddct = rq = None
        if ct_t is not None and ct_r is not None:
            dct = ct_t - ct_r
            if ave_blank is not None:
                ddct = dct - ave_blank
                rq = 2 ** (-ddct)
                rq_by_sample.setdefault(r["sample"], []).append(rq)
        enriched.append((r, ct_t, ct_r, dct, ddct, rq))

    rq_stats: dict[str, tuple[str, str]] = {}
    for s, vals in rq_by_sample.items():
        mean = f"{round(statistics.mean(vals), 4)}" if vals else ""
        sd = f"{round(statistics.stdev(vals), 4)}" if len(vals) > 1 else ("0.0" if vals else "")
        rq_stats[s] = (mean, sd)

    # 列名对齐报告 xlsx(图三)：内参基因/目的基因(存CT值)、Ave. Blank、Mean、SD；
    # 位置拆「目的孔」「内参孔」两列；去掉「样本」列(仍按样本内部分组算 Mean/SD)。
    header = [
        "目的孔", "内参孔", "内参基因", "目的基因",
        "ΔCT", "Ave. Blank", "ΔΔCT", "2^(-ΔΔCT)", "Mean", "SD",
    ]
    ave_blank_str = _num(ave_blank)  # 仅整表首行显值(对齐图三 =AVERAGE 只挂首行)
    rows: list[list[str]] = []
    prev_sample = None
    for idx, (r, ct_t, ct_r, dct, ddct, rq) in enumerate(enriched):
        sample = r["sample"]
        if sample != prev_sample:
            mean, sd = rq_stats.get(sample, ("", ""))
            prev_sample = sample
        else:
            mean = sd = ""  # 同一样本组仅首行显示均值/SD
        rows.append([
            r["pos_target"],
            r["pos_ref"],
            _num(ct_r) if ct_r is not None else "Undetermined",
            _num(ct_t) if ct_t is not None else "Undetermined",
            _num(dct),
            ave_blank_str if idx == 0 else "",
            _num(ddct),
            _num(rq),
            mean,
            sd,
        ])
    return {
        "header": header,
        "rows": rows,
        "meta": {
            "source": os.path.basename(xml_path),
            "threshold": threshold,
            "ave_blank": round(ave_blank, 4) if ave_blank is not None else None,
        },
    }


# ===========================================================================
# main
# ===========================================================================

def parse_args():
    p = argparse.ArgumentParser(
        description="场景二实验报告一键生成(RNA浓度 + qPCR板排布/曲线/统计 -> 单个 xlsx)",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--fluor-csv", default=DEFAULT_FLUOR, help="BioTek 荧光 CSV(RNA 浓度)")
    p.add_argument("--xml", default=DEFAULT_XML, help="qPCR 原始数据 XML")
    p.add_argument("--plate-map", default=DEFAULT_MAP, help="96 孔细胞培养板 xlsx(两部分共用)")
    p.add_argument("--out", default=DEFAULT_OUT, help="输出 xlsx 路径")
    p.add_argument("--slope", type=float, default=1.0, help="RNA 标曲斜率(占位)")
    p.add_argument("--intercept", type=float, default=0.0, help="RNA 标曲截距(占位)")
    p.add_argument("--blank-mode", choices=["subtract", "none"], default="subtract",
                   help="RNA 是否扣除空白均值")
    p.add_argument("--threshold", type=float, default=3.0, help="qPCR CT 阈值(原始荧光)")
    p.add_argument("--amp-cycles", type=int, default=45, help="扩增段循环数")
    p.add_argument("--max-points", type=int, default=None,
                   help="每孔最多取多少采集点; 默认全部(扩增+熔解)")
    p.add_argument("--logy", action="store_true", help="扩增曲线纵轴取对数")
    p.add_argument("--dpi", type=int, default=150, help="图片 dpi")
    return p.parse_args()


def build_report_xlsx(
    fluor_csv: str,
    xml_path: str,
    out_xlsx: str,
    plate_map_path: str = DEFAULT_MAP,
    slope: float = 1.0,
    intercept: float = 0.0,
    blank_mode: str = "subtract",
    threshold: float = 3.0,
    amp_cycles: int = 45,
    max_points: int | None = None,
    dpi: int = 150,
    logy: bool = False,
) -> str:
    """装配场景二完整报告工作簿(RNA浓度 / 板排布+扩增曲线 / qPCR统计与分析)并保存。

    供节点 m(compute_experiment2_result) 直接调用，生成可下载附件 xlsx；
    ``main()`` 亦复用本函数。返回 ``out_xlsx`` 路径。
    """
    out_dir = os.path.dirname(os.path.abspath(out_xlsx))
    os.makedirs(out_dir, exist_ok=True)
    # 中间 PNG 落到独立子目录，避免覆盖调用方(节点)同目录已上传的 qpcr_*.png
    img_dir = os.path.join(out_dir, "_report_imgs")
    os.makedirs(img_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    default_ws = wb.active  # 移除默认空 sheet(后面都用 create_sheet)

    # ---- 第 1 部分: RNA 浓度 ----
    conc_by_sample, rna_meta, blank_mean = compute_rna(
        fluor_csv, plate_map_path, slope, intercept, blank_mode)
    add_rna_sheet(wb, conc_by_sample, fluor_csv, rna_meta, blank_mean, slope, intercept)

    # ---- 第 2 部分: 解析 qPCR + 映射 ----
    pos_by_sample, fluor_by_sample = parse_qpcr_xml(xml_path, max_points)
    npoints = max((len(v) for v in fluor_by_sample.values()), default=0)
    fluor_by_pos = {pos_by_sample[sn]: vals for sn, vals in fluor_by_sample.items()
                    if sn in pos_by_sample}
    map96 = read_plate_map(plate_map_path)
    meta = build_well_meta(map96)

    # ---- 绘图 + 板排布/曲线 sheet ----
    layout_png = os.path.join(img_dir, "qpcr_plate_layout.png")
    curve_png = os.path.join(img_dir, "qpcr_amp_curves.png")
    plot_plate_layout(meta, map96, layout_png, dpi)
    plot_amp_curves(fluor_by_pos, meta, curve_png, dpi, amp_cycles=amp_cycles, logy=logy)
    add_plate_sheets(wb, layout_png, curve_png, fluor_by_pos, meta, npoints)

    # ---- 第 3 部分: 统计与分析 ----
    ct_by_pos = {pos: ct_threshold(vals[:amp_cycles], threshold)
                 for pos, vals in fluor_by_pos.items()}
    rows = build_rows(map96, ct_by_pos)
    ave_blank = compute_ave_blank(rows)
    add_stats_sheet(wb, rows, xml_path, threshold, ave_blank)

    # 删除默认空 sheet 并保存(图片随本次一次性写入, 不会丢失)
    wb.remove(default_ws)
    wb.save(out_xlsx)
    return out_xlsx


def main() -> int:
    args = parse_args()
    print(f"[*] RNA={args.fluor_csv}\n    XML={args.xml}\n    MAP={args.plate_map}")
    build_report_xlsx(
        args.fluor_csv,
        args.xml,
        args.out,
        plate_map_path=args.plate_map,
        slope=args.slope,
        intercept=args.intercept,
        blank_mode=args.blank_mode,
        threshold=args.threshold,
        amp_cycles=args.amp_cycles,
        max_points=args.max_points,
        dpi=args.dpi,
        logy=args.logy,
    )
    print(f"完成 -> {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
