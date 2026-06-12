import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 96 孔板：横向 1-12，纵向 A-H
COLS = list(range(1, 13))
ROWS = [chr(ord("A") + i) for i in range(8)]  # A-H
TOTAL = len(COLS) * len(ROWS)  # 96

# 样品数量
SAMPLE_COUNT = 11
# 每种内容至少出现的次数，其余随机填满
MIN_EACH = 4

# 需要随机填充的内容：空白对照 + 样品1-ID ~ 样品N-ID
LABELS = ["空白对照"] + [f"样品{i}-ID" for i in range(1, SAMPLE_COUNT + 1)]

# 先保证每种至少 MIN_EACH 个，剩余位置随机填充
assert MIN_EACH * len(LABELS) <= TOTAL, "MIN_EACH 太大，放不下 96 个孔"
values = LABELS * MIN_EACH
values += [random.choice(LABELS) for _ in range(TOTAL - len(values))]
random.shuffle(values)

wb = Workbook()
ws = wb.active
ws.title = "细胞培养板"

# 样式
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(color="FFFFFF", bold=True)
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 左上角标题
top_left = ws.cell(row=1, column=1, value="细胞培养板ID")
top_left.fill = header_fill
top_left.font = header_font
top_left.alignment = center
top_left.border = border

# 顶部列号 1-12
for j, c in enumerate(COLS, start=2):
    cell = ws.cell(row=1, column=j, value=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# 行标题 A-H + 随机填充内容
idx = 0
for i, r in enumerate(ROWS, start=2):
    rc = ws.cell(row=i, column=1, value=r)
    rc.fill = header_fill
    rc.font = header_font
    rc.alignment = center
    rc.border = border
    for j in range(2, len(COLS) + 2):
        cell = ws.cell(row=i, column=j, value=values[idx])
        cell.alignment = center
        cell.border = border
        idx += 1

# 列宽
ws.column_dimensions["A"].width = 14
for j in range(2, len(COLS) + 2):
    ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 12

out = "实验二细胞培养板.xlsx"
wb.save(out)

# 统计每种数量，方便核对
from collections import Counter

print("已生成:", out)
for label, n in sorted(Counter(values).items()):
    print(f"  {label}: {n}")
