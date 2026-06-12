"""生成实验一孔板图 实验一细胞培养板.xlsx（最多 16 块 96 孔板，支持 4 通量）。

仿 data/gen_plate.py，但一次产出多个 sheet：细胞培养板1 .. 细胞培养板16，
每个 sheet 是一块 96 孔板(8x12)，单元格随机填 空白对照 / 样品N-ID。
sheet 名与提交期落盘文件 m 中的板名(细胞培养板{seq})一一对应，
作为 compute_experiment1_result 计算时「孔位 -> 样本身份」的来源。

多通量：每通量 4 块板，最多 4 通量共 16 块。若已存在 实验一细胞培养板.xlsx，
则保留其已有 sheet，仅补建缺失的 细胞培养板{n}（不破坏既有布局）。

用法:
    python3 data/gen_experiment1_plate.py
"""

import os
import random
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 96 孔板：横向 1-12，纵向 A-H
COLS = list(range(1, 13))
ROWS = [chr(ord("A") + i) for i in range(8)]  # A-H
TOTAL = len(COLS) * len(ROWS)  # 96

# 板数量与样品数量（16 块 = 最多 4 通量 × 每通量 4 块）
PLATE_COUNT = 16
SAMPLE_COUNT = 11
# 每种内容至少出现的次数，其余随机填满
MIN_EACH = 4

# 需要随机填充的内容：空白对照 + 样品1-ID ~ 样品N-ID
LABELS = ["空白对照"] + [f"样品{i}-ID" for i in range(1, SAMPLE_COUNT + 1)]
assert MIN_EACH * len(LABELS) <= TOTAL, "MIN_EACH 太大，放不下 96 个孔"

# 样式
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(color="FFFFFF", bold=True)
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill_sheet(ws, plate_no: int) -> Counter:
    """在 ws 上画一块 96 孔板，返回各标签计数。"""
    # 先保证每种至少 MIN_EACH 个，剩余随机填充
    values = LABELS * MIN_EACH
    values += [random.choice(LABELS) for _ in range(TOTAL - len(values))]
    random.shuffle(values)

    # 左上角标题
    top_left = ws.cell(row=1, column=1, value=f"细胞培养板{plate_no}ID")
    top_left.fill = header_fill
    top_left.font = header_font
    top_left.alignment = center
    top_left.border = border

    # 顶部列号 1-12
    for j, c in enumerate(COLS, start=2):
        cell = ws.cell(row=1, column=j, value=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 行标题 A-H + 填充内容
    idx = 0
    for i, r in enumerate(ROWS, start=2):
        rc = ws.cell(row=i, column=1, value=r)
        rc.fill = header_fill
        rc.font = header_font
        rc.alignment = center
        rc.border = border
        for j in range(2, len(COLS) + 2):
            cell = ws.cell(row=i, column=j, value=values[idx])
            cell.alignment = center
            cell.border = border
            idx += 1

    # 列宽
    ws.column_dimensions["A"].width = 14
    for j in range(2, len(COLS) + 2):
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 12

    return Counter(values)


def main() -> None:
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "实验一细胞培养板.xlsx")

    # 已存在则保留既有 sheet，仅补建缺失的；否则新建
    if os.path.exists(out):
        wb = load_workbook(out)
        default_ws = None
        print("已加载现有文件，保留既有 sheet:", wb.sheetnames)
    else:
        wb = Workbook()
        default_ws = wb.active

    for plate_no in range(1, PLATE_COUNT + 1):
        title = f"细胞培养板{plate_no}"
        if title in wb.sheetnames:
            print(f"{title}: 已存在，跳过")
            continue
        ws = wb.create_sheet(title=title)
        counts = fill_sheet(ws, plate_no)
        print(f"{title}:")
        for label, n in sorted(counts.items()):
            print(f"  {label}: {n}")

    if default_ws is not None:
        wb.remove(default_ws)
    wb.save(out)
    print("已生成:", out, "共 sheet:", len(wb.sheetnames))


if __name__ == "__main__":
    main()
